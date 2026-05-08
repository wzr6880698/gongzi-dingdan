"""
生产报工工资匹配计划单和订单分析工具 (Streamlit 网页版)
"""

import streamlit as st
import pandas as pd
import numpy as np
import os
import re
import io
import tempfile
from fuzzywuzzy import fuzz, process
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

# ============================================================
# 页面配置
# ============================================================
st.set_page_config(
    page_title="生产报工工资匹配计划单和订单分析工具",
    page_icon="🐾",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ============================================================
# 自定义 CSS 样式
# ============================================================
st.markdown("""
<style>
    .block-container { padding-top: 2rem; }
    .stFileUploader { border: 2px dashed #D4875A; border-radius: 8px; padding: 1rem; }
    .stSuccess { background-color: #D4EDDA; padding: 0.5rem 1rem; border-radius: 4px; }
    .stWarning { background-color: #FFF3CD; padding: 0.5rem 1rem; border-radius: 4px; }
    .stError { background-color: #F8D7DA; padding: 0.5rem 1rem; border-radius: 4px; }
    div[data-testid="stSidebar"] { background-color: #F8F0EB; }
    .section-header {
        font-size: 1.2rem;
        font-weight: 600;
        color: #D4875A;
        border-bottom: 2px solid #D4875A;
        padding-bottom: 0.3rem;
        margin-top: 1.5rem;
        margin-bottom: 0.8rem;
    }
</style>
""", unsafe_allow_html=True)


# ============================================================
# 1. 智能探测表头行号
# ============================================================
def detect_header_row(file_obj, sheet_name=0, scan_rows=20, keywords=None):
    """智能探测表头所在行号"""
    if keywords is None:
        keywords = ['订单', '产品', '数量', '单价', '金额', '工价', '工资', '姓名', '规格']

    df_raw = pd.read_excel(file_obj, sheet_name=sheet_name, header=None, nrows=scan_rows)
    best_row, best_score = 0, 0
    for idx, row in df_raw.iterrows():
        row_str = ' '.join([str(cell).lower() for cell in row if pd.notna(cell)])
        score = sum(1 for kw in keywords if kw in row_str)
        if score > best_score:
            best_score, best_row = score, idx

    if best_score < 2:
        st.warning(f"工作表 [{sheet_name}] 自动探测表头失败，将默认使用第1行作为表头。")
        return 0
    else:
        st.success(f"工作表 [{sheet_name}] 表头位于第 {best_row + 1} 行（得分 {best_score}）")
        return best_row


def read_sheet_auto_header(file_obj, sheet_name=0):
    """自动探测表头并读取工作表"""
    header_row = detect_header_row(file_obj, sheet_name)
    df = pd.read_excel(file_obj, sheet_name=sheet_name, header=header_row)
    df = df.dropna(how='all').dropna(axis=1, how='all')
    df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
    return df


def read_all_sheets_auto_header(file_obj):
    """遍历所有工作表自动读取并合并"""
    xl = pd.ExcelFile(file_obj)
    sheet_names = xl.sheet_names
    st.info(f"发现 {len(sheet_names)} 个工作表: {sheet_names}")

    dfs = []
    for sheet in sheet_names:
        try:
            with st.spinner(f"正在读取工作表: {sheet}"):
                df_sheet = read_sheet_auto_header(file_obj, sheet)
            if df_sheet.empty:
                st.warning(f"工作表 [{sheet}] 无有效数据，已跳过")
                continue
            df_sheet['_来源工作表'] = sheet
            dfs.append(df_sheet)
        except Exception as e:
            st.error(f"读取工作表 [{sheet}] 失败: {e}")

    if not dfs:
        raise ValueError("所有工作表均无有效数据")

    merged_df = pd.concat(dfs, ignore_index=True, sort=False)
    st.success(f"所有工作表合并完成，共 {len(merged_df)} 行数据。")
    return merged_df


# ============================================================
# 2. 数据清洗函数
# ============================================================
def clean_text(x):
    """清洗文本：去除多余空格"""
    if pd.isna(x):
        return ''
    s = str(x).strip()
    s = re.sub(r'\s+', ' ', s)
    return s


def clean_order_id(x):
    """清理订单号：仅去除首尾空格和不可见控制字符，保留原始所有字符"""
    if pd.isna(x):
        return ''
    s = str(x).strip()
    s = re.sub(r'[\n\r\t]', '', s)
    return s


def clean_numeric(x):
    """清洗数值：去除货币符号、千分位逗号等"""
    if pd.isna(x):
        return 0.0
    s = str(x).strip()
    if s == '':
        return 0.0
    s = re.sub(r'[¥￥$€\s]', '', s)
    s = re.sub(r'(?<=\d),(?=\d{3})', '', s)
    try:
        return float(s)
    except ValueError:
        s = re.sub(r'[^\d\.\-]', '', s)
        try:
            return float(s)
        except:
            return 0.0


def combine_name_spec(name, spec):
    """合并产品名称和规格"""
    name = clean_text(name)
    spec = clean_text(spec) if spec else ''
    if spec:
        return f"{name} | {spec}"
    else:
        return name


# ============================================================
# 3. 列名识别（增强版）
# ============================================================
def find_column(df, keywords, allow_multiple=False, description="", strict_mode=False):
    """在df的列名中查找包含指定关键词的列"""
    matches = []
    for kw in keywords:
        kw_lower = kw.lower()
        for col in df.columns:
            col_lower = str(col).lower()
            if kw_lower in col_lower:
                matches.append(col)
                break
        if matches and not allow_multiple:
            break

    if not matches:
        if description:
            st.warning(f"未找到列：{description}，关键词 {keywords}")
        return None

    if description:
        st.info(f"识别到 {description} 列: `{matches[0]}` (关键词 '{keywords[0]}' 匹配)")
    return matches if allow_multiple else matches[0]


def find_customer_column(df, description="客户名称列"):
    """
    智能识别客户名称列，避免匹配到“客户编码”等列。
    策略：
      1. 优先查找列名完全等于“客户”或“客户名称”的列
      2. 其次查找列名包含“客户”且不包含“编码”、“code”的列
      3. 最后回退到包含“客户”的第一个列
    """
    cols_lower = {col: str(col).lower() for col in df.columns}

    # 1. 完全匹配
    for col, col_lower in cols_lower.items():
        if col_lower in ['客户', '客户名称', '客户名']:
            st.info(f"识别到 {description}: `{col}` (完全匹配)")
            return col

    # 2. 包含“客户”但不含“编码”“code”
    for col, col_lower in cols_lower.items():
        if '客户' in col_lower:
            if '编码' not in col_lower and 'code' not in col_lower:
                st.info(f"识别到 {description}: `{col}` (包含'客户'且不含'编码')")
                return col

    # 3. 包含“客户”的任意列（最后手段）
    for col, col_lower in cols_lower.items():
        if '客户' in col_lower:
            st.warning(f"识别到 {description}: `{col}` (可能包含编码信息，请确认)")
            return col

    st.warning(f"未找到 {description}，将填充默认值'未知'")
    return None


# ============================================================
# 4. 订单内产品匹配
# ============================================================
def match_product_within_order(labor_key, order_id, sales_dict, threshold=75):
    """在订单内进行产品名称模糊匹配"""
    candidates = sales_dict.get(order_id, [])
    if not candidates:
        return None, None

    labor_clean = clean_text(labor_key)
    for cand in candidates:
        if clean_text(cand) == labor_clean:
            return cand, cand

    if len(candidates) == 1:
        return candidates[0], candidates[0]

    match, score = process.extractOne(labor_clean, candidates, scorer=fuzz.token_sort_ratio)
    if score >= threshold:
        st.info(f"映射: 订单[{order_id}]内 '{labor_key}' -> '{match}' (相似度 {score})")
        return match, match
    else:
        st.warning(f"订单[{order_id}]内 '{labor_key}' 未能匹配，候选产品: {candidates} (最高相似度 {score})")
        return None, None


# ============================================================
# 5. 保存 Excel 并合并单元格，增加高亮预警（支持单行高亮）
# ============================================================
def save_with_merge_and_highlight(df, merge_cols, sum_cols, highlight_col, threshold=20.0):
    """
    保存 DataFrame 为 Excel，合并指定列，并对 highlight_col 中大于 threshold 的单元格标红。
    支持多行订单合并及单行订单高亮。
    """
    buffer = io.BytesIO()
    df.to_excel(buffer, index=False, engine='openpyxl')
    buffer.seek(0)

    wb = load_workbook(buffer)
    ws = wb.active

    start_row = 2  # 数据起始行（第1行是表头）
    total_rows = len(df) + 1

    # 按销售订单号分组
    groups = {}
    current_order = None
    group_start = start_row
    for i, order in enumerate(df['销售订单号'], start=start_row):
        if order != current_order:
            if current_order is not None:
                groups[current_order] = (group_start, i - 1)
            current_order = order
            group_start = i
    if current_order is not None:
        groups[current_order] = (group_start, total_rows)

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for order, (g_start, g_end) in groups.items():
        # ---------- 合并操作（仅多行订单需要合并）----------
        if g_start != g_end:
            # 合并文本列（例如“销售订单号”及“客户名称”等）
            for col_name in merge_cols:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    ws.merge_cells(start_row=g_start, start_column=col_idx,
                                   end_row=g_end, end_column=col_idx)
                    cell = ws.cell(row=g_start, column=col_idx)
                    cell.alignment = Alignment(vertical='center')

            # 合并汇总列（数值列，需要保留合计值）
            for col_name in sum_cols:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1
                    # 清除合并区域下方单元格的值
                    for r in range(g_start + 1, g_end + 1):
                        ws.cell(row=r, column=col_idx).value = None
                    ws.merge_cells(start_row=g_start, start_column=col_idx,
                                   end_row=g_end, end_column=col_idx)
                    cell = ws.cell(row=g_start, column=col_idx)
                    cell.alignment = Alignment(vertical='center')

        # ---------- 高亮处理（所有订单：多行或单行都执行）----------
        if highlight_col in df.columns:
            col_idx = df.columns.get_loc(highlight_col) + 1
            # 获取该订单对应高亮列的单元格（多行订单已合并，左上角单元格即为合并后单元格）
            cell = ws.cell(row=g_start, column=col_idx)
            val_str = cell.value
            numeric_val = 0.0

            # 将单元格内容转换为数值（支持百分比字符串、数字）
            if isinstance(val_str, (int, float)):
                numeric_val = float(val_str)
            elif isinstance(val_str, str):
                cleaned = val_str.strip().replace('%', '')
                try:
                    numeric_val = float(cleaned)
                except:
                    numeric_val = 0.0
            else:
                numeric_val = 0.0

            if numeric_val > threshold:
                # 对整个订单区域（g_start 到 g_end）的所有行的高亮列填充红色
                for row in range(g_start, g_end + 1):
                    ws.cell(row=row, column=col_idx).fill = red_fill

    output_buffer = io.BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer


# ============================================================
# 6. 核心处理逻辑
# ============================================================
def process_data(sales_file, prod_file, labor_file, match_threshold=75, highlight_threshold=20.0):
    """核心数据处理流程，返回结果DataFrame、输出Excel字节流、总工资、总销售额"""

    # ---------- 读取文件 ----------
    with st.status("读取销售订单表...", expanded=True) as status:
        df_sales = read_sheet_auto_header(sales_file, 0)
        status.update(label="销售订单表读取完成", state="complete")

    with st.status("读取生产指令单...", expanded=True) as status:
        df_prod = read_sheet_auto_header(prod_file, 0)
        status.update(label="生产指令单读取完成", state="complete")

    with st.status("读取车间报工表（遍历所有工作表）...", expanded=True) as status:
        df_labor = read_all_sheets_auto_header(labor_file)
        status.update(label="车间报工表读取完成", state="complete")

    # ---------- 调试信息 ----------
    with st.expander("调试信息：销售订单表列名", expanded=False):
        st.write(list(df_sales.columns))

    # ---------- 识别列名 ----------
    st.markdown('<div class="section-header">列名自动识别</div>', unsafe_allow_html=True)

    # 销售表列名
    col_order_sales = find_column(df_sales, ['单据编号', '订单号', '订单编号', '销售订单', '单据号'],
                                  description="销售订单号列")
    col_product_sales = find_column(df_sales, ['产品', '产品名称', '商品', '存货', '存货名称'],
                                    description="销售产品名列")
    col_spec_sales = find_column(df_sales, ['规格', '规格型号', '型号', '存货规格'], description="规格型号列")
    col_qty_sales = find_column(df_sales, ['数量', '销售数量'], description="销售数量列")
    col_price = find_column(df_sales, ['本币含税单价', '单价', '销售单价', '含税单价'], description="销售单价列")
    col_amount = find_column(df_sales, ['本币含税金额', '金额', '销售金额', '总价', '含税金额'],
                             description="销售金额列")
    # 使用智能函数识别客户名称列
    col_customer = find_customer_column(df_sales, description="客户名称列")

    if not all([col_order_sales, col_product_sales, col_qty_sales, col_price, col_amount]):
        st.error("销售订单表缺少必要列，请检查文件后重新上传。")
        return None, None, None, None, None
    spec_col_sales = col_spec_sales if col_spec_sales else None

    # 生产指令单列名
    col_order_prod = find_column(df_prod, ['订单号', '销售订单号', '订单编号', '单据编号'], description="生产订单号列")
    col_product_prod = find_column(df_prod, ['产品名称', '生产产品', '产品', '存货'], description="生产产品名列")
    col_spec_prod = find_column(df_prod, ['规格', '规格型号', '型号', '产品规格'], description="生产规格列")
    col_plan_qty = find_column(df_prod, ['生产数量', '计划数量', '数量'], description="计划数量列")

    if not all([col_order_prod, col_product_prod, col_plan_qty]):
        st.error("生产指令单表缺少必要列，请检查文件后重新上传。")
        return None, None, None, None, None
    spec_col_prod = col_spec_prod if col_spec_prod else None

    # 报工表列名
    col_order_labor = find_column(df_labor, ['订单号', '销售订单号', '订单编号', '所属销售订单'],
                                  description="报工订单号列")
    col_product_labor = find_column(df_labor, ['产品名称', '生产产品', '产品', '存货'], description="报工产品名列")
    col_spec_labor = find_column(df_labor, ['规格', '规格型号', '型号', '产品规格'], description="报工规格列")
    col_actual_qty = find_column(df_labor, ['数量', '实际数量', '生产数量', '数量/工时'], description="实际数量列")
    col_wage_amount = find_column(df_labor, ['薪资金额', '工资金额', '金额', '工资', '工价'], description="薪资金额列")

    if not all([col_order_labor, col_product_labor, col_actual_qty, col_wage_amount]):
        st.error("报工表缺少必要列，请检查文件后重新上传。")
        return None, None, None, None, None
    spec_col_labor = col_spec_labor if col_spec_labor else None

    # ---------- 统一列名 ----------
    rename_sales = {
        col_order_sales: '销售订单号',
        col_product_sales: '销售产品名',
        col_qty_sales: '销售数量',
        col_price: '销售单价',
        col_amount: '销售金额'
    }
    if spec_col_sales:
        rename_sales[spec_col_sales] = '规格型号'
    if col_customer:
        rename_sales[col_customer] = '客户名称'
    df_sales = df_sales.rename(columns=rename_sales)

    rename_prod = {
        col_order_prod: '销售订单号',
        col_product_prod: '生产产品名',
        col_plan_qty: '计划生产数量'
    }
    if spec_col_prod:
        rename_prod[spec_col_prod] = '规格型号'
    df_prod = df_prod.rename(columns=rename_prod)

    rename_labor = {
        col_order_labor: '销售订单号',
        col_product_labor: '报工产品名',
        col_actual_qty: '实际生产数量',
        col_wage_amount: '薪资金额'
    }
    if spec_col_labor:
        rename_labor[spec_col_labor] = '规格型号'
    df_labor = df_labor.rename(columns=rename_labor)

    # 补充缺失的规格列
    for df in [df_sales, df_prod, df_labor]:
        if '规格型号' not in df.columns:
            df['规格型号'] = ''

    # 如果销售表中没有客户名称列，则创建一个空列并给出提示
    if '客户名称' not in df_sales.columns:
        df_sales['客户名称'] = '未知'
        st.info("销售订单表中未发现客户名称列，已自动填充为'未知'。")

    # ---------- 数据清洗 ----------
    st.markdown('<div class="section-header">数据清洗</div>', unsafe_allow_html=True)

    with st.spinner("正在清洗文本数据..."):
        for df in [df_sales, df_prod, df_labor]:
            df['销售订单号'] = df['销售订单号'].apply(clean_order_id)
        df_sales['销售产品名'] = df_sales['销售产品名'].apply(clean_text)
        df_sales['规格型号'] = df_sales['规格型号'].apply(clean_text)
        df_sales['客户名称'] = df_sales['客户名称'].apply(clean_text)
        df_prod['生产产品名'] = df_prod['生产产品名'].apply(clean_text)
        df_prod['规格型号'] = df_prod['规格型号'].apply(clean_text)
        df_labor['报工产品名'] = df_labor['报工产品名'].apply(clean_text)
        df_labor['规格型号'] = df_labor['规格型号'].apply(clean_text)
    st.success("文本清洗完成")

    # 创建产品键
    df_sales['产品键'] = df_sales.apply(lambda r: combine_name_spec(r['销售产品名'], r['规格型号']), axis=1)
    df_prod['产品键'] = df_prod.apply(lambda r: combine_name_spec(r['生产产品名'], r['规格型号']), axis=1)
    df_labor['产品键'] = df_labor.apply(lambda r: combine_name_spec(r['报工产品名'], r['规格型号']), axis=1)

    # 数值清洗
    with st.spinner("正在清洗数值数据..."):
        df_sales['销售数量'] = df_sales['销售数量'].apply(clean_numeric)
        df_sales['销售单价'] = df_sales['销售单价'].apply(clean_numeric)
        df_sales['销售金额'] = df_sales['销售金额'].apply(clean_numeric)
        df_prod['计划生产数量'] = df_prod['计划生产数量'].apply(clean_numeric)
        df_labor['实际生产数量'] = df_labor['实际生产数量'].apply(clean_numeric)
        df_labor['薪资金额'] = df_labor['薪资金额'].apply(clean_numeric)
    st.success("数值清洗完成")

    # 过滤空订单号
    df_sales = df_sales[df_sales['销售订单号'] != '']
    df_prod = df_prod[df_prod['销售订单号'] != '']
    df_labor = df_labor[df_labor['销售订单号'] != '']

    # ---------- 构建客户名称映射 ----------
    customer_map = df_sales.groupby('销售订单号')['客户名称'].first().to_dict()

    # ---------- 订单级汇总 ----------
    st.markdown('<div class="section-header">订单级汇总计算</div>', unsafe_allow_html=True)

    order_sales_sum = df_sales.groupby('销售订单号').agg(
        按订单统计数量=('销售数量', 'sum'),
        按订单统计金额=('销售金额', 'sum')
    ).reset_index()

    order_wage_sum = df_labor.groupby('销售订单号').agg(
        按订单统计生产工资额=('薪资金额', 'sum')
    ).reset_index()

    # 构建销售字典
    sales_dict = df_sales.groupby('销售订单号')['产品键'].unique().apply(list).to_dict()

    # 报工明细汇总
    labor_agg = df_labor.groupby(['销售订单号', '产品键', '报工产品名', '规格型号'], as_index=False).agg({
        '实际生产数量': 'sum',
        '薪资金额': 'sum'
    })

    # ---------- 产品匹配 ----------
    st.markdown('<div class="section-header">产品名称匹配</div>', unsafe_allow_html=True)

    match_log = []
    matched_keys = []
    for _, row in labor_agg.iterrows():
        order = row['销售订单号']
        labor_key = row['产品键']
        matched_key, _ = match_product_within_order(labor_key, order, sales_dict, threshold=match_threshold)
        matched_keys.append(matched_key)
        match_log.append({
            '订单号': order,
            '报工产品': labor_key,
            '匹配结果': matched_key if matched_key else '未匹配'
        })
    labor_agg['映射产品键'] = matched_keys
    labor_agg['映射产品键'] = labor_agg['映射产品键'].fillna(labor_agg['产品键'])

    # 显示匹配日志
    match_df = pd.DataFrame(match_log)
    unmatched = match_df[match_df['匹配结果'] == '未匹配']
    if not unmatched.empty:
        st.warning(f"共有 {len(unmatched)} 条产品未能自动匹配，请检查数据。")
        with st.expander("查看未匹配记录", expanded=False):
            st.dataframe(unmatched, use_container_width=True)
    else:
        st.success("所有产品均已成功匹配！")

    # 生产计划汇总
    prod_agg = df_prod.groupby(['销售订单号', '产品键', '生产产品名', '规格型号'], as_index=False).agg({
        '计划生产数量': 'sum'
    })

    # 销售明细汇总
    sales_agg = df_sales.groupby(['销售订单号', '产品键', '销售产品名', '规格型号'], as_index=False).agg({
        '销售数量': 'sum',
        '销售单价': 'first',
        '销售金额': 'sum'
    })

    # ---------- 合并明细 ----------
    st.markdown('<div class="section-header">数据合并与计算</div>', unsafe_allow_html=True)

    with st.spinner("正在合并销售、生产、报工数据..."):
        merged = labor_agg.merge(prod_agg, on=['销售订单号', '产品键'], how='left', suffixes=('', '_prod'))
        merged = merged.merge(sales_agg, on=['销售订单号', '产品键'], how='left', suffixes=('', '_sales'))

        merged['销售产品名'] = merged['销售产品名'].fillna(merged['报工产品名'])
        merged['规格型号'] = merged['规格型号'].fillna(
            merged['规格型号_labor'] if '规格型号_labor' in merged.columns else ''
        )
        merged['计划生产数量'] = merged['计划生产数量'].fillna(0)
        merged['销售数量'] = merged['销售数量'].fillna(0)
        merged['销售金额'] = merged['销售金额'].fillna(0)
        merged['销售单价'] = merged['销售单价'].fillna(0)
        # 若销售金额为空但数量单价有值，计算补充
        merged['销售金额'] = merged.apply(
            lambda row: row['销售数量'] * row['销售单价']
            if row['销售金额'] == 0 and row['销售数量'] > 0
            else row['销售金额'],
            axis=1
        )

        # 合并订单汇总数据
        merged = merged.merge(order_sales_sum, on='销售订单号', how='left')
        merged = merged.merge(order_wage_sum, on='销售订单号', how='left')

        merged['按订单统计数量'] = merged['按订单统计数量'].fillna(0)
        merged['按订单统计金额'] = merged['按订单统计金额'].fillna(0)
        merged['按订单统计生产工资额'] = merged['按订单统计生产工资额'].fillna(0)

    st.success("数据合并完成")

    # ---------- 计算工资占比 ----------
    with st.spinner("正在计算工资占比..."):
        merged['工资额占订单销售额比'] = merged.apply(
            lambda row: (row['按订单统计生产工资额'] / row['按订单统计金额'] * 100)
            if row['按订单统计金额'] != 0 else 0.0,
            axis=1
        )
        merged['工资额占订单销售额比'] = merged['工资额占订单销售额比'].apply(lambda x: f"{x:.2f}%")

    # ---------- 最终结果 ----------
    result_cols = [
        '销售订单号', '销售产品名', '规格型号',
        '销售数量', '销售单价', '销售金额',
        '计划生产数量', '实际生产数量', '薪资金额',
        '按订单统计数量', '按订单统计金额', '按订单统计生产工资额',
        '工资额占订单销售额比'
    ]
    result = merged[result_cols].copy()
    result.rename(columns={'薪资金额': '生产工资额'}, inplace=True)

    # ---------- 插入客户名称列（位于“销售订单号”之后）----------
    order_col_idx = result.columns.get_loc('销售订单号')
    result.insert(order_col_idx + 1, '客户名称', result['销售订单号'].map(customer_map))
    result['客户名称'] = result['客户名称'].fillna('未知')

    result = result.sort_values(['销售订单号', '销售产品名', '规格型号']).reset_index(drop=True)

    total_wage = result['生产工资额'].sum()
    total_sales = result['按订单统计金额'].drop_duplicates().sum()

    # ---------- 预警统计（需要临时数值列）----------
    result['_temp_ratio'] = result['工资额占订单销售额比'].str.rstrip('%').astype(float)
    warning_orders_df = result[result['_temp_ratio'] > highlight_threshold][
        ['销售订单号', '客户名称', '销售产品名', '按订单统计金额', '按订单统计生产工资额', '工资额占订单销售额比']]
    # 删除临时列（确保最终 Excel 中不包含该列）
    result.drop(columns=['_temp_ratio'], inplace=True)

    # ---------- 生成 Excel ----------
    with st.spinner("正在生成 Excel 文件（含合并单元格和高亮预警）..."):
        merge_cols = ['销售订单号', '客户名称']  # 合并这两列
        sum_cols = ['按订单统计数量', '按订单统计金额', '按订单统计生产工资额', '工资额占订单销售额比']
        highlight_col = '工资额占订单销售额比'
        excel_buffer = save_with_merge_and_highlight(
            result, merge_cols, sum_cols, highlight_col, threshold=highlight_threshold
        )

    st.success("Excel 文件生成完成！")

    return result, excel_buffer, total_wage, total_sales, warning_orders_df


# ============================================================
# 7. Streamlit 主界面
# ============================================================
def main():
    # ---------- 标题区域 ----------
    st.title("生产工资与计划和订单汇总统计分析工具")
    st.caption("自动读取销售订单、生产指令单、车间报工台账，智能匹配产品并计算工资占比预警")

    st.markdown("---")

    # ---------- 侧边栏参数配置 ----------
    with st.sidebar:
        st.header("参数配置")

        match_threshold = st.slider(
            "产品名称匹配阈值",
            min_value=50,
            max_value=100,
            value=75,
            step=5,
            help="模糊匹配相似度阈值，越高要求越严格"
        )

        highlight_threshold = st.slider(
            "工资占比预警阈值 (%)",
            min_value=5,
            max_value=50,
            value=20,
            step=5,
            help="超过此阈值的订单将在 Excel 中以红色底色高亮"
        )

        st.markdown("---")
        st.header("使用说明")
        st.markdown("""
        **步骤 1**：上传三个 Excel 文件
        - 销售订单台账（需包含客户名称列）
        - 生产指令单台账
        - 车间报工台账

        **步骤 2**：点击「开始处理」按钮

        **步骤 3**：查看结果并下载 Excel

        **注意事项**：
        - 系统会自动探测表头行
        - 支持报工表多工作表合并
        - 产品名称支持模糊匹配
        - 工资占比超阈值会红色高亮（包括单行订单）
        - 输出 Excel 包含客户名称列，并自动合并相同订单的行
        - **客户名称优先匹配“客户”或“客户名称”列，避免误取“客户编码”**
        """)

    # ---------- 文件上传区域 ----------
    st.markdown('<div class="section-header">上传数据文件</div>', unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)

    with col1:
        sales_file = st.file_uploader(
            "销售订单台账",
            type=['xlsx', 'xls'],
            key="sales_uploader",
            help="上传销售订单台账 Excel 文件"
        )
        if sales_file:
            st.success(f"已选择: {sales_file.name}")

    with col2:
        prod_file = st.file_uploader(
            "生产指令单台账",
            type=['xlsx', 'xls'],
            key="prod_uploader",
            help="上传生产指令单台账 Excel 文件"
        )
        if prod_file:
            st.success(f"已选择: {prod_file.name}")

    with col3:
        labor_file = st.file_uploader(
            "车间报工台账",
            type=['xlsx', 'xls'],
            key="labor_uploader",
            help="上传车间报工台账 Excel 文件（支持多工作表）"
        )
        if labor_file:
            st.success(f"已选择: {labor_file.name}")

    st.markdown("---")

    # ---------- 处理按钮 ----------
    all_files_ready = all([sales_file, prod_file, labor_file])

    if not all_files_ready:
        st.info("请上传全部三个文件后，点击下方按钮开始处理。")
        uploaded = sum([1 for f in [sales_file, prod_file, labor_file] if f])
        st.progress(uploaded / 3, text=f"文件上传进度: {uploaded}/3")

    process_btn = st.button(
        "开始处理",
        type="primary",
        disabled=not all_files_ready,
        use_container_width=True
    )

    # ---------- 执行处理 ----------
    if process_btn and all_files_ready:
        try:
            result, excel_buffer, total_wage, total_sales, warning_orders_df = process_data(
                sales_file, prod_file, labor_file,
                match_threshold=match_threshold,
                highlight_threshold=highlight_threshold
            )

            if result is not None and excel_buffer is not None:
                # ---------- 汇总统计卡片 ----------
                st.markdown('<div class="section-header">汇总统计</div>', unsafe_allow_html=True)

                kpi1, kpi2, kpi3, kpi4 = st.columns(4)
                with kpi1:
                    st.metric("订单总数", f"{result['销售订单号'].nunique()}")
                with kpi2:
                    st.metric("产品明细行数", f"{len(result)}")
                with kpi3:
                    st.metric("全月生产工资合计", f"¥{total_wage:,.2f}")
                with kpi4:
                    avg_ratio = (total_wage / total_sales * 100) if total_sales > 0 else 0
                    st.metric("平均工资占销售比", f"{avg_ratio:.2f}%")

                # ---------- 预警订单 ----------
                if not warning_orders_df.empty:
                    st.warning(
                        f"发现 {warning_orders_df['销售订单号'].nunique()} 个订单的工资占比超过 {highlight_threshold}% 预警阈值！")
                    with st.expander("查看预警订单详情", expanded=False):
                        st.dataframe(
                            warning_orders_df,
                            use_container_width=True,
                            hide_index=True
                        )

                # ---------- 数据预览 ----------
                st.markdown('<div class="section-header">数据预览</div>', unsafe_allow_html=True)

                tab_preview, tab_all = st.tabs(["前20行预览", "完整数据"])
                with tab_preview:
                    st.dataframe(result.head(20), use_container_width=True, hide_index=True)
                with tab_all:
                    st.dataframe(result, use_container_width=True, hide_index=True)

                # ---------- 下载按钮 ----------
                st.markdown('<div class="section-header">下载结果</div>', unsafe_allow_html=True)

                st.download_button(
                    label="下载汇总结果 Excel（含客户名称、合并单元格和红色预警高亮）",
                    data=excel_buffer,
                    file_name="生产工资匹配订单汇总统计表.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )

                st.balloons()

        except Exception as e:
            st.error(f"处理过程中发生错误: {e}")
            st.exception(e)


if __name__ == "__main__":
    main()